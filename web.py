#!/usr/bin/env python3
"""
自选股实时行情 - 完整版
支持：实时行情 | 添加/删除 | 批量导入 | 编辑
"""

import json
import http.server
import socketserver
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse, parse_qs
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

PORT = 8892
DATA_DIR = Path(__file__).parent / "data"
WATCHLIST_FILE = DATA_DIR / "watchlist.json"

DATA_DIR.mkdir(exist_ok=True)

DEFAULT_STOCKS = [
    {"code": "600519", "name": "贵州茅台", "market": "sh"},
    {"code": "000858", "name": "五粮液", "market": "sz"},
    {"code": "300750", "name": "宁德时代", "market": "sz"},
    {"code": "000001", "name": "平安银行", "market": "sz"},
    {"code": "601318", "name": "中国平安", "market": "sh"},
    {"code": "000333", "name": "美的集团", "market": "sz"},
    {"code": "600036", "name": "招商银行", "market": "sh"},
    {"code": "601888", "name": "中国中免", "market": "sh"},
]

def load_watchlist():
    if not WATCHLIST_FILE.exists():
        with open(WATCHLIST_FILE, 'w', encoding='utf-8') as f:
            json.dump(DEFAULT_STOCKS, f, ensure_ascii=False, indent=2)
        return DEFAULT_STOCKS
    with open(WATCHLIST_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_watchlist(stocks):
    with open(WATCHLIST_FILE, 'w', encoding='utf-8') as f:
        json.dump(stocks, f, ensure_ascii=False, indent=2)

def get_realtime():
    """获取实时行情"""
    watchlist = load_watchlist()
    if not watchlist:
        return []
    
    try:
        symbols = [f"{s['market']}{s['code']}" for s in watchlist]
        url = f"http://qt.gtimg.cn/q={','.join(symbols)}"
        resp = requests.get(url, timeout=10)
        resp.encoding = 'gbk'
        
        stocks = []
        lines = resp.text.strip().split('\n')
        
        for i, line in enumerate(lines):
            if '="="' in line or not line.strip():
                continue
            parts = line.split('~')
            if len(parts) < 40:
                continue
            
            stock = watchlist[i] if i < len(watchlist) else {}
            price = float(parts[3]) if parts[3].replace('.', '').isdigit() else 0
            prev_close = float(parts[4]) if parts[4].replace('.', '').isdigit() else 0
            change = price - prev_close if prev_close > 0 else 0
            change_pct = (change / prev_close * 100) if prev_close > 0 else 0
            
            stocks.append({
                'code': parts[2] if len(parts) > 2 else stock.get('code', ''),
                'name': stock.get('name', parts[1] if len(parts) > 1 else ''),
                'market': stock.get('market', ''),
                'price': price,
                'prev_close': prev_close,
                'change': round(change, 2),
                'change_pct': round(change_pct, 2),
                'open': float(parts[5]) if parts[5].replace('.', '').isdigit() else 0,
                'high': float(parts[33]) if parts[33].replace('.', '').isdigit() else 0,
                'low': float(parts[34]) if parts[34].replace('.', '').isdigit() else 0,
                'volume': float(parts[36]) if parts[36].replace('.', '').isdigit() else 0,
                'amount': float(parts[37]) if parts[37].replace('.', '').isdigit() else 0,
            })
        return stocks
    except Exception as e:
        print(f"获取行情失败: {e}")
        return []

def get_stock_name(code):
    """根据代码获取股票名称"""
    market = 'sh' if code.startswith(('6', '5', '9')) else 'sz'
    try:
        url = f"http://qt.gtimg.cn/q={market}{code}"
        resp = requests.get(url, timeout=5)
        resp.encoding = 'gbk'
        parts = resp.text.split('~')
        if len(parts) > 1:
            return parts[1]
    except:
        pass
    return code

def get_kline_data(code, market, period='day', count=100):
    """获取K线数据"""
    try:
        if period == 'day' or period == '5' or period == '10' or period == 'week' or period == 'month' or period == 'year':
            url = f"https://money.finance.sina.com.cn/quotes_service/api/json_v2.php/CN_MarketData.getKLineData"
            params = {
                'symbol': f"{market}{code}",
                'scale': period if period != 'day' else 240,
                'ma': 'no',
                'datalen': count
            }
            resp = requests.get(url, params=params, timeout=10)
            if resp.text and resp.text != 'null':
                return resp.json()
        elif period == 'fen':
            url = f"https://quotes.sina.cn/cn/api/openapi.php/StockChartService.getMinutelyData"
            params = {'symbol': f"{market}{code}"}
            resp = requests.get(url, params=params, timeout=10)
            if resp.text:
                data = resp.json()
                if data.get('result') and data['result'].get('data'):
                    items = data['result']['data'].get('day', [])
                    return [{'day': item[0], 'close': float(item[1]), 'open': float(item[2]), 
                             'high': float(item[3]), 'low': float(item[4]), 'volume': float(item[5])} 
                            for item in items[-count:]]
    except Exception as e:
        print(f"获取K线失败 {code}: {e}")
    return []

def calc_rsi(prices, period=14):
    """计算RSI"""
    if len(prices) < period + 1:
        return None
    deltas = [prices[i] - prices[i-1] for i in range(1, len(prices))]
    gains = [d if d > 0 else 0 for d in deltas]
    losses = [-d if d < 0 else 0 for d in deltas]
    
    avg_gain = sum(gains[-period:]) / period
    avg_loss = sum(losses[-period:]) / period
    
    if avg_loss == 0:
        return 100
    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))

def calc_macd(prices, fast=12, slow=26, signal=9):
    """计算MACD，返回(DIF, DEA, MACD)"""
    if len(prices) < slow + signal:
        return None, None, None
    
    # 计算EMA
    def ema(data, n):
        k = 2 / (n + 1)
        result = [data[0]]
        for i in range(1, len(data)):
            result.append(data[i] * k + result[-1] * (1 - k))
        return result
    
    ema_fast = ema(prices, fast)
    ema_slow = ema(prices, slow)
    dif = [ema_fast[i] - ema_slow[i] for i in range(len(ema_fast))]
    dea = ema(dif, signal)
    macd = [(dif[i] - dea[i]) * 2 for i in range(len(dif))]
    
    return dif[-1], dea[-1], macd[-1]

def calc_atr(kdata, period=14):
    """计算ATR"""
    if len(kdata) < period + 1:
        return None
    
    true_ranges = []
    for i in range(1, len(kdata)):
        high = float(kdata[i].get('high', 0))
        low = float(kdata[i].get('low', 0))
        prev_close = float(kdata[i-1].get('close', 0))
        
        tr = max(high - low, abs(high - prev_close), abs(low - prev_close))
        true_ranges.append(tr)
    
    if len(true_ranges) < period:
        return None
    return sum(true_ranges[-period:]) / period

def get_market_sentiment():
    """获取市场情绪数据"""
    try:
        # 东方财富行情API
        url = "https://push2.eastmoney.com/api/qt/stock/get"
        params = {
            'ut': 'fa5fd1943c7b386f172d6893dbfba10b',
            'fltt': 2,
            'invt': 2,
            'secid': '1.000001',  # 上证指数
            'fields': 'f43,f44,f45,f46,f47,f48,f57,f58'
        }
        resp = requests.get(url, params=params, timeout=5)
        data = resp.json()
        if data.get('data'):
            d = data['data']
            return {
                'shanghai_up': '获取中',  # 需要另外的API
                'shanghai_down': '获取中',
                'limit_up': '获取中',
                'limit_down': '获取中',
                'main_line': 'AI/科技',
                'risk_level': '中等'
            }
    except:
        pass
    return {
        'shanghai_up': '-',
        'shanghai_down': '-', 
        'limit_up': '-',
        'limit_down': '-',
        'main_line': '待确认',
        'risk_level': '待评估'
    }

def export_excel():
    """导出专业短线交易观察池Excel"""
    watchlist = load_watchlist()
    stocks = get_realtime()
    
    if not stocks:
        return None
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    
    # 样式定义
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
    header_fill_gold = PatternFill(start_color="B7950B", end_color="B7950B", fill_type="solid")
    center = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin = Side(style='thin', color='666666')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    red_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    gold_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
    blue_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    
    # ====== Sheet1: 观察池 ======
    ws = wb.active
    ws.title = "观察池"
    
    headers = ['股票代码', '股票名称', '主线方向', '市场地位', '最新价', '涨跌幅', 
               '成交额', '换手率', '量比', '5日涨幅', '20日涨幅', '接近新高', '强度评分', '备注']
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    
    row_num = 2
    for s in stocks:
        code = s['code']
        name = s['name']
        price = s['price']
        change_pct = s.get('change_pct', 0)
        amount = s.get('amount', 0)
        market = s.get('market', 'sz')
        
        # 获取额外数据
        try:
            url = f"http://qt.gtimg.cn/q={market}{code}"
            resp = requests.get(url, timeout=5)
            resp.encoding = 'gbk'
            parts = resp.text.split('~')
            turnover = float(parts[38]) if len(parts) > 38 and parts[38].replace('.', '').isdigit() else 0
            vol_ratio = float(parts[49]) if len(parts) > 49 and parts[49].replace('.', '').isdigit() else 0
        except:
            turnover = 0
            vol_ratio = 0
        
        # 获取K线数据计算指标
        kdata = get_kline_data(code, market, 'day', 100)
        closes = [float(d.get('close', 0)) for d in kdata] if kdata else []
        
        # 5日涨幅
        change_5d = 0
        if len(closes) >= 5:
            change_5d = (closes[-1] - closes[-5]) / closes[-5] * 100 if closes[-5] > 0 else 0
        
        # 20日涨幅
        change_20d = 0
        if len(closes) >= 20:
            change_20d = (closes[-1] - closes[-20]) / closes[-20] * 100 if closes[-20] > 0 else 0
        
        # 接近历史新高
        near_high = "否"
        if len(closes) >= 60:
            high_60d = max(closes[-60:])
            if closes[-1] >= high_60d * 0.97:  # 97%以上视为接近
                near_high = "是"
        
        # 计算强度评分（0-100）
        score = 50  # 基础分
        if change_pct > 0:
            score += min(change_pct * 3, 20)  # 涨幅加分
        else:
            score += max(change_pct * 3, -20)  # 跌幅减分
        score += min(change_5d / 2, 15)  # 5日强势加分
        score += min(change_20d / 4, 15)  # 20日趋势加分
        if near_high == "是":
            score += 10  # 接近新高加分
        if vol_ratio > 1.5:
            score += 5  # 量比高加分
        score = max(0, min(100, int(score)))  # 限制在0-100
        
        # 格式化数据
        change_str = f"{'+' if change_pct > 0 else ''}{change_pct:.2f}%"
        change_5d_str = f"{'+' if change_5d > 0 else ''}{change_5d:.2f}%"
        change_20d_str = f"{'+' if change_20d > 0 else ''}{change_20d:.2f}%"
        
        if amount >= 100000000:
            amount_str = f"{amount/100000000:.2f}亿"
        elif amount >= 10000:
            amount_str = f"{amount/10000:.2f}万"
        else:
            amount_str = f"{amount:.0f}"
        
        # 自动判断主线方向（可优化）
        direction = ""
        if any(x in name for x in ['光', '通信', 'CPO']):
            direction = "光模块"
        elif any(x in name for x in ['算力', 'AI', '智能']):
            direction = "AI算力"
        elif any(x in name for x in ['电', '网', '电气']):
            direction = "电网"
        elif any(x in name for x in ['新能', '锂', '宁德']):
            direction = "新能源"
        elif any(x in name for x in ['半导', '芯', '微']):
            direction = "半导体"
        
        # 自动判断市场地位
        position = ""
        if score >= 80:
            position = "龙头"
        elif score >= 65:
            position = "中军"
        elif score >= 50:
            position = "趋势核心"
        elif score >= 35:
            position = "补涨"
        else:
            position = "跟风"
        
        row_data = [
            code, name, direction, position, price, change_str,
            amount_str,
            f"{turnover:.2f}%" if turnover > 0 else "-",
            f"{vol_ratio:.2f}" if vol_ratio > 0 else "-",
            change_5d_str, change_20d_str, near_high, f"{score}分", ""
        ]
        
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row_num, col, val)
            cell.border = border
            cell.alignment = center
        
        # 颜色标记
        if change_pct > 0:
            for col in range(1, 15):
                ws.cell(row_num, col).fill = red_fill
        elif change_pct < 0:
            for col in range(1, 15):
                ws.cell(row_num, col).fill = green_fill
        
        # 强度评分颜色
        if score >= 80:
            ws.cell(row_num, 13).fill = gold_fill
        elif score < 30:
            ws.cell(row_num, 13).fill = blue_fill
        
        row_num += 1
    
    # 设置列宽
    col_widths = [12, 12, 10, 10, 10, 10, 10, 8, 8, 10, 10, 8, 10, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    
    ws.freeze_panes = 'A2'
    
    # ====== Sheet2: 日K ======
    ws_k = wb.create_sheet(title="日K")
    
    k_headers = ['股票代码', '日期', '开盘', '最高', '最低', '收盘', 
                 '成交量', 'MA5', 'MA10', 'MA20', 'MA60', 'RSI', 'MACD', 'ATR']
    
    for col, h in enumerate(k_headers, 1):
        cell = ws_k.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    
    row_num = 2
    for s in stocks:
        code = s['code']
        market = s.get('market', 'sz')
        kdata = get_kline_data(code, market, 'day', 100)
        
        if kdata:
            closes = [float(d.get('close', 0)) for d in kdata]
            highs = [float(d.get('high', 0)) for d in kdata]
            lows = [float(d.get('low', 0)) for d in kdata]
            volumes = [float(d.get('volume', 0)) for d in kdata]
            
            for i, d in enumerate(kdata):
                close = closes[i]
                
                # 计算MA
                ma5 = sum(closes[max(0,i-4):i+1]) / min(i+1, 5) if i >= 0 else 0
                ma10 = sum(closes[max(0,i-9):i+1]) / min(i+1, 10) if i >= 0 else 0
                ma20 = sum(closes[max(0,i-19):i+1]) / min(i+1, 20) if i >= 0 else 0
                ma60 = sum(closes[max(0,i-59):i+1]) / min(i+1, 60) if i >= 0 else 0
                
                # 计算RSI
                rsi = calc_rsi(closes[:i+1], 14) if i >= 14 else None
                
                # 计算MACD
                dif, dea, macd = calc_macd(closes[:i+1])
                
                # 计算ATR
                atr = calc_atr(kdata[:i+1], 14) if i >= 14 else None
                
                row_data = [
                    code,
                    d.get('day', ''),
                    float(d.get('open', 0)),
                    float(d.get('high', 0)),
                    float(d.get('low', 0)),
                    close,
                    volumes[i] if i < len(volumes) else 0,
                    round(ma5, 2) if ma5 > 0 else '',
                    round(ma10, 2) if ma10 > 0 else '',
                    round(ma20, 2) if ma20 > 0 else '',
                    round(ma60, 2) if ma60 > 0 else '',
                    f"{rsi:.1f}" if rsi else '',
                    f"{macd:.2f}" if macd else '',
                    f"{atr:.2f}" if atr else ''
                ]
                
                for col, val in enumerate(row_data, 1):
                    cell = ws_k.cell(row_num, col, val)
                    cell.border = border
                    cell.alignment = center
                
                # 颜色标记
                if i > 0:
                    if close >= closes[i-1]:
                        for col in range(1, 15):
                            ws_k.cell(row_num, col).fill = red_fill
                    else:
                        for col in range(1, 15):
                            ws_k.cell(row_num, col).fill = green_fill
                
                row_num += 1
    
    # 设置日K列宽
    for col in 'ABCDEFG':
        ws_k.column_dimensions[col].width = 12
    for col in 'HIJKLMN':
        ws_k.column_dimensions[col].width = 10
    
    ws_k.freeze_panes = 'A2'
    
    # ====== Sheet3: 持仓 ======
    ws_pos = wb.create_sheet(title="持仓")
    
    pos_headers = ['股票名称', '持仓股数', '可卖股数', '成本价', '当前价', 
                   '浮盈亏', '浮盈亏率', '仓位占比', '今日买入']
    
    for col, h in enumerate(pos_headers, 1):
        cell = ws_pos.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    
    # 加载持仓数据
    pos_file = DATA_DIR / "positions.json"
    positions = []
    if pos_file.exists():
        with open(pos_file, 'r', encoding='utf-8') as f:
            positions = json.load(f)
    
    # 合并持仓数据
    row_num = 2
    for pos in positions:
        stock_name = pos.get('name', '')
        shares = pos.get('shares', 0)
        cost = pos.get('cost', 0)
        
        # 找当前价
        current_price = 0
        for s in stocks:
            if s['name'] == stock_name:
                current_price = s['price']
                break
        
        # 计算浮盈亏
        if shares > 0 and cost > 0:
            profit = (current_price - cost) * shares
            profit_pct = (current_price - cost) / cost * 100 if cost > 0 else 0
        else:
            profit = 0
            profit_pct = 0
        
        row_data = [
            stock_name, shares, shares, cost, current_price,
            f"{'+' if profit > 0 else ''}{profit:.2f}",
            f"{'+' if profit_pct > 0 else ''}{profit_pct:.2f}%",
            "-",
            0
        ]
        
        for col, val in enumerate(row_data, 1):
            cell = ws_pos.cell(row_num, col, val)
            cell.border = border
            cell.alignment = center
        
        # 颜色
        if profit > 0:
            for col in range(1, 10):
                ws_pos.cell(row_num, col).fill = red_fill
        elif profit < 0:
            for col in range(1, 10):
                ws_pos.cell(row_num, col).fill = green_fill
        
        row_num += 1
    
    # 设置持仓列宽
    for col in 'ABCDE':
        ws_pos.column_dimensions[col].width = 12
    for col in 'FGHI':
        ws_pos.column_dimensions[col].width = 10
    
    ws_pos.freeze_panes = 'A2'
    
    # ====== Sheet4: 市场情绪 ======
    ws_sent = wb.create_sheet(title="市场情绪")
    
    sentiment_headers = ['指标名称', '数值', '说明']
    
    for col, h in enumerate(sentiment_headers, 1):
        cell = ws_sent.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill_gold
        cell.alignment = center
        cell.border = border
    
    # 获取市场情绪
    sent = get_market_sentiment()
    
    # 计算涨跌家数
    up_count = sum(1 for s in stocks if s.get('change_pct', 0) > 0)
    down_count = sum(1 for s in stocks if s.get('change_pct', 0) < 0)
    avg_change = sum(s.get('change_pct', 0) for s in stocks) / len(stocks) if stocks else 0
    
    sentiment_data = [
        ('上涨家数', up_count, '观察池中上涨股票数'),
        ('下跌家数', down_count, '观察池中下跌股票数'),
        ('平盘数量', len(stocks) - up_count - down_count, '涨跌幅为0'),
        ('平均涨跌幅', f"{'+' if avg_change > 0 else ''}{avg_change:.2f}%", '算术平均'),
        ('最强股票', f"{max(stocks, key=lambda x: x.get('change_pct', 0))['name']}" if stocks else '-', '今日涨幅最大'),
        ('最弱股票', f"{min(stocks, key=lambda x: x.get('change_pct', 0))['name']}" if stocks else '-', '今日跌幅最大'),
        ('当前主线', sent.get('main_line', '待确认'), '当日最强板块'),
        ('市场风险', sent.get('risk_level', '待评估'), '综合评分'),
        ('更新时间', datetime.now().strftime('%Y-%m-%d %H:%M'), '数据时间'),
    ]
    
    for i, (name, value, desc) in enumerate(sentiment_data, 2):
        ws_sent.cell(i, 1, name).border = border
        ws_sent.cell(i, 1).alignment = center
        ws_sent.cell(i, 2, value).border = border
        ws_sent.cell(i, 2).alignment = center
        ws_sent.cell(i, 3, desc).border = border
        ws_sent.cell(i, 3).alignment = left_align
        ws_sent.cell(i, 3).font = Font(color="888888", size=10)
    
    ws_sent.column_dimensions['A'].width = 15
    ws_sent.column_dimensions['B'].width = 20
    ws_sent.column_dimensions['C'].width = 25
    
    # 保存
    filename = f"观察池_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = DATA_DIR / filename
    wb.save(filepath)
    return filename

HTML = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>自选股行情</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { 
            font-family: -apple-system, BlinkMacSystemFont, 'PingFang SC', sans-serif;
            background: #1a1a2e; color: #fff; min-height: 100vh; padding-bottom: 80px;
        }
        .header {
            background: linear-gradient(135deg, #1a1a2e, #16213e);
            padding: 15px; position: sticky; top: 0; z-index: 100;
            border-bottom: 1px solid #333;
        }
        .header-top { display: flex; justify-content: space-between; align-items: center; }
        .header h1 { font-size: 18px; color: #ffd700; }
        .header-btns { display: flex; gap: 8px; }
        .btn {
            background: #0f3460; color: #fff; border: none;
            padding: 8px 15px; border-radius: 5px; cursor: pointer; font-size: 14px;
        }
        .btn:active { background: #1a5276; }
        .btn-primary { background: #e94560; }
        .time { color: #888; font-size: 12px; margin-top: 8px; }
        
        /* 自选股标签栏 */
        .watchlist-bar {
            background: #16213e; padding: 10px; display: flex; gap: 8px;
            overflow-x: auto; border-bottom: 1px solid #333;
        }
        .watchlist-bar::-webkit-scrollbar { display: none; }
        .tag {
            background: #0f3460; padding: 5px 12px; border-radius: 15px;
            font-size: 12px; white-space: nowrap; cursor: pointer;
        }
        .tag .del { color: #e94560; margin-left: 5px; }
        .tag:hover { background: #1a5276; }
        
        /* 股票列表 */
        .stock-list { padding: 10px; }
        .stock-card {
            background: linear-gradient(135deg, #16213e, #1a1a2e);
            border-radius: 12px; margin-bottom: 10px; padding: 15px;
            border: 1px solid #333;
        }
        .stock-header { display: flex; justify-content: space-between; align-items: center; }
        .stock-name { font-size: 16px; font-weight: bold; }
        .stock-code { font-size: 12px; color: #888; }
        .stock-price { font-size: 24px; font-weight: bold; }
        .up { color: #f23645; }
        .down { color: #089981; }
        .stock-info {
            display: grid; grid-template-columns: repeat(3, 1fr); gap: 8px;
            font-size: 12px; color: #aaa; margin-top: 10px;
        }
        .info-item { text-align: center; background: #0f3460; padding: 8px; border-radius: 5px; }
        .info-label { color: #666; margin-bottom: 2px; }
        
        /* 模态框 */
        .modal {
            display: none; position: fixed; top: 0; left: 0; right: 0; bottom: 0;
            background: rgba(0,0,0,0.8); z-index: 1000; padding: 20px;
        }
        .modal.active { display: flex; align-items: center; justify-content: center; }
        .modal-content {
            background: #16213e; border-radius: 12px; padding: 20px;
            width: 100%; max-width: 400px; max-height: 80vh; overflow-y: auto;
        }
        .modal-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 15px; }
        .modal-title { font-size: 18px; color: #ffd700; }
        .modal-close { background: none; border: none; color: #888; font-size: 24px; cursor: pointer; }
        
        /* 表单 */
        .form-group { margin-bottom: 15px; }
        .form-label { display: block; margin-bottom: 5px; color: #888; font-size: 14px; }
        .form-input {
            width: 100%; padding: 12px; border-radius: 8px; border: 1px solid #333;
            background: #1a1a2e; color: #fff; font-size: 14px;
        }
        .form-input:focus { outline: none; border-color: #ffd700; }
        .form-hint { font-size: 12px; color: #666; margin-top: 5px; }
        
        /* 编辑列表 */
        .edit-list { max-height: 300px; overflow-y: auto; }
        .edit-item {
            display: flex; justify-content: space-between; align-items: center;
            padding: 10px; background: #1a1a2e; border-radius: 8px; margin-bottom: 8px;
        }
        .edit-item-info { flex: 1; }
        .edit-item-name { font-weight: bold; }
        .edit-item-code { font-size: 12px; color: #888; }
        .edit-item-del { color: #e94560; background: none; border: none; font-size: 18px; cursor: pointer; }
        
        .loading, .error { text-align: center; padding: 40px; color: #888; }
    </style>
</head>
<body>
    <!-- 顶部导航 -->
    <div class="header">
        <div class="header-top">
            <h1>📈 自选股</h1>
            <div class="header-btns">
                <button class="btn" onclick="openEdit()">✏️ 编辑</button>
                <button class="btn btn-primary" onclick="openAdd()">➕ 添加</button>
                <button class="btn" onclick="exportData()" id="exportBtn">📥 导出</button>
            </div>
        </div>
        <div class="time" id="updateTime">--</div>
        
        <!-- 自选股标签 -->
        <div class="watchlist-bar" id="watchlistBar">
            <span style="color:#666;font-size:12px;">暂无自选股</span>
        </div>
    </div>
    
    <!-- 股票列表 -->
    <div class="stock-list" id="stockList">
        <div class="loading">加载中...</div>
    </div>

    <!-- 添加股票模态框 -->
    <div class="modal" id="addModal">
        <div class="modal-content">
            <div class="modal-header">
                <div class="modal-title">添加股票</div>
                <button class="modal-close" onclick="closeAdd()">×</button>
            </div>
            <div class="form-group">
                <label class="form-label">股票代码</label>
                <input type="text" class="form-input" id="inputCode" placeholder="如: 600519">
                <div class="form-hint">支持6位股票代码，自动识别沪/深</div>
            </div>
            <div class="form-group">
                <label class="form-label">股票名称</label>
                <input type="text" class="form-input" id="inputName" placeholder="如: 贵州茅台">
            </div>
            <div class="form-group">
                <label class="form-label">批量导入 (每行一个)</label>
                <textarea class="form-input" id="inputBatch" rows="5" placeholder="600519,贵州茅台
000858,五粮液
300750,宁德时代"></textarea>
                <div class="form-hint">格式: 代码,名称 (用逗号分隔)</div>
            </div>
            <button class="btn btn-primary" style="width:100%" onclick="doAdd()">确认添加</button>
        </div>
    </div>

    <!-- 编辑模态框 -->
    <div class="modal" id="editModal">
        <div class="modal-content">
            <div class="modal-header">
                <div class="modal-title">编辑自选股</div>
                <button class="modal-close" onclick="closeEdit()">×</button>
            </div>
            <div class="edit-list" id="editList"></div>
            <button class="btn" style="width:100%;margin-top:15px" onclick="closeEdit()">完成</button>
        </div>
    </div>

    <script>
    let stocks = [];
    let watchlist = [];
    
    // 加载数据
    function loadData() {
        document.getElementById('stockList').innerHTML = '<div class="loading">加载中...</div>';
        Promise.all([
            fetch('/api/quote').then(r => r.json()),
            fetch('/api/list').then(r => r.json())
        ]).then(([q, w]) => {
            stocks = q.stocks || [];
            watchlist = w || [];
            renderStockList();
            renderWatchlist();
            document.getElementById('updateTime').textContent = '更新: ' + q.time;
        }).catch(e => {
            document.getElementById('stockList').innerHTML = '<div class="error">加载失败</div>';
        });
    }
    
    // 渲染股票列表
    function renderStockList() {
        const el = document.getElementById('stockList');
        if (!stocks.length) {
            el.innerHTML = '<div class="error">暂无数据，点击右上角添加</div>';
            return;
        }
        
        el.innerHTML = stocks.map(s => {
            const cls = s.change > 0 ? 'up' : s.change < 0 ? 'down' : '';
            const arrow = s.change > 0 ? '▲' : s.change < 0 ? '▼' : '─';
            
            return `<div class="stock-card">
                <div class="stock-header">
                    <div>
                        <div class="stock-name">${s.name}</div>
                        <div class="stock-code">${s.code}</div>
                    </div>
                    <div style="text-align:right">
                        <div class="stock-price ${cls}">${s.price}</div>
                        <div class="${cls}">${arrow} ${Math.abs(s.change)} (${arrow} ${Math.abs(s.change_pct)}%)</div>
                    </div>
                </div>
                <div class="stock-info">
                    <div class="info-item">
                        <div class="info-label">今开</div>
                        <div>${s.open}</div>
                    </div>
                    <div class="info-item">
                        <div class="info-label">最高</div>
                        <div class="up">${s.high}</div>
                    </div>
                    <div class="info-item">
                        <div class="info-label">最低</div>
                        <div class="down">${s.low}</div>
                    </div>
                    <div class="info-item">
                        <div class="info-label">成交量</div>
                        <div>${fmt(s.volume)}</div>
                    </div>
                    <div class="info-item">
                        <div class="info-label">成交额</div>
                        <div>${fmtAmt(s.amount)}</div>
                    </div>
                    <div class="info-item">
                        <div class="info-label">昨收</div>
                        <div>${s.prev_close}</div>
                    </div>
                </div>
            </div>`;
        }).join('');
    }
    
    // 渲染标签栏
    function renderWatchlist() {
        const el = document.getElementById('watchlistBar');
        if (!watchlist.length) {
            el.innerHTML = '<span style="color:#666;font-size:12px;">暂无自选股</span>';
            return;
        }
        el.innerHTML = watchlist.map(s => 
            `<span class="tag">${s.code} ${s.name}</span>`
        ).join('');
    }
    
    // 格式化数字
    function fmt(n) {
        if (n >= 100000000) return (n/100000000).toFixed(2) + '亿';
        if (n >= 10000) return (n/10000).toFixed(2) + '万';
        return n.toFixed(0);
    }
    function fmtAmt(n) {
        if (n >= 100000000) return (n/100000000).toFixed(2) + '亿';
        if (n >= 10000) return (n/10000).toFixed(2) + '万';
        return n.toFixed(0);
    }
    
    // 添加模态框
    function openAdd() {
        document.getElementById('addModal').classList.add('active');
        document.getElementById('inputCode').value = '';
        document.getElementById('inputName').value = '';
        document.getElementById('inputBatch').value = '';
    }
    function closeAdd() {
        document.getElementById('addModal').classList.remove('active');
    }
    
    // 编辑模态框
    function openEdit() {
        document.getElementById('editModal').classList.add('active');
        renderEditList();
    }
    function closeEdit() {
        document.getElementById('editModal').classList.remove('active');
        loadData();
    }
    
    function renderEditList() {
        const el = document.getElementById('editList');
        if (!watchlist.length) {
            el.innerHTML = '<div style="text-align:center;color:#888;padding:20px;">暂无自选股</div>';
            return;
        }
        el.innerHTML = watchlist.map(s => `
            <div class="edit-item">
                <div class="edit-item-info">
                    <div class="edit-item-name">${s.name}</div>
                    <div class="edit-item-code">${s.code} (${s.market === 'sh' ? '沪' : '深'})</div>
                </div>
                <button class="edit-item-del" onclick="deleteStock('${s.code}')">×</button>
            </div>
        `).join('');
    }
    
    // 添加股票
    async function doAdd() {
        const code = document.getElementById('inputCode').value.trim();
        const name = document.getElementById('inputName').value.trim();
        const batch = document.getElementById('inputBatch').value.trim();
        
        let added = 0;
        
        // 添加单个
        if (code && name) {
            const r = await fetch(`/api/add?code=${code}&name=${encodeURIComponent(name)}`);
            const d = await r.json();
            if (d.success) added++;
        }
        
        // 批量添加
        if (batch) {
            const lines = batch.split('\\n');
            for (const line of lines) {
                const parts = line.split(',');
                if (parts.length >= 2) {
                    const c = parts[0].trim();
                    const n = parts[1].trim();
                    if (c && n) {
                        await fetch(`/api/add?code=${c}&name=${encodeURIComponent(n)}`);
                        added++;
                    }
                }
            }
        }
        
        if (added > 0) {
            closeAdd();
            loadData();
        } else {
            alert('请输入股票代码和名称');
        }
    }
    
    // 删除股票
    async function deleteStock(code) {
        if (!confirm('确定删除?')) return;
        await fetch(`/api/remove?code=${code}`);
        renderEditList();
    }
    
    // 导出Excel
    async function exportData() {
        const btn = document.getElementById('exportBtn');
        btn.textContent = '导出中...';
        btn.disabled = true;
        
        try {
            // 先调用导出API生成文件
            const r = await fetch('/api/export');
            const d = await r.json();
            if (d.success) {
                // 直接下载最新的文件
                window.location.href = '/download/export';
                btn.textContent = '✅ 已导出';
                setTimeout(() => { btn.textContent = '📥 导出'; }, 3000);
            } else {
                alert(d.error || '导出失败');
                btn.textContent = '📥 导出';
            }
        } catch(e) {
            alert('导出失败');
            btn.textContent = '📥 导出';
        }
        btn.disabled = false;
    }
    
    // 页面加载
    window.onload = loadData;
    setInterval(loadData, 30000);
    </script>
</body>
</html>'''

class Handler(http.server.SimpleHTTPRequestHandler):
    def send_json(self, data):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))
    
    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        
        if path == '/' or path == '/index.html':
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(HTML.encode('utf-8'))
        elif path == '/api/quote':
            self.send_json({'stocks': get_realtime(), 'time': datetime.now().strftime('%H:%M:%S')})
        elif path == '/api/list':
            self.send_json(load_watchlist())
        elif path == '/api/add':
            params = parse_qs(parsed.query)
            code = params.get('code', [''])[0]
            name = params.get('name', [''])[0]
            market = 'sh' if code.startswith(('6', '5', '9')) else 'sz'
            
            stocks = load_watchlist()
            if not any(s['code'] == code for s in stocks):
                stocks.append({'code': code, 'name': name, 'market': market})
                save_watchlist(stocks)
                self.send_json({'success': True})
            else:
                self.send_json({'success': False, 'error': '已存在'})
        elif path == '/api/remove':
            params = parse_qs(parsed.query)
            code = params.get('code', [''])[0]
            
            stocks = load_watchlist()
            stocks = [s for s in stocks if s['code'] != code]
            save_watchlist(stocks)
            self.send_json({'success': True})
        elif path == '/api/export':
            filename = export_excel()
            if filename:
                self.send_json({'success': True, 'filename': filename})
            else:
                self.send_json({'success': False, 'error': '导出失败'})
        elif path.startswith('/download/'):
            import glob
            suffix = path[10:]
            # 根据suffix查找对应文件
            if suffix == 'export':
                pattern = str(DATA_DIR / "观察池_*.xlsx")
            elif suffix == 'latest':
                pattern = str(DATA_DIR / "观察池_*.xlsx")
            else:
                pattern = str(DATA_DIR / f"*{suffix}*.xlsx")
            
            matches = sorted(glob.glob(pattern), reverse=True)
            
            if matches:
                filepath = Path(matches[0])
                filename = filepath.name
            else:
                filepath = None
                filename = None
            
            if filepath and filepath.exists():
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                # 使用RFC 5987编码支持非ASCII文件名
                from urllib.parse import quote
                encoded_filename = quote(filename)
                self.send_header('Content-Disposition', f"attachment; filename*=UTF-8''{encoded_filename}")
                self.end_headers()
                with open(filepath, 'rb') as f:
                    self.wfile.write(f.read())
            else:
                self.send_error(404)
        else:
            super().do_GET()

if __name__ == '__main__':
    print(f"╔═══════════════════════════════════════════╗")
    print(f"║  自选股行情 Web {PORT}                      ║")
    print(f"╠═══════════════════════════════════════════╣")
    print(f"║  本机: http://localhost:{PORT}              ║")
    print(f"║  局域网: http://192.168.2.78:{PORT}       ║")
    print(f"╚═══════════════════════════════════════════╝")
    
    socketserver.TCPServer.allow_reuse_address = True
    with socketserver.TCPServer(("0.0.0.0", PORT), Handler) as httpd:
        httpd.serve_forever()